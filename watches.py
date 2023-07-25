from bs4 import BeautifulSoup
from openpyxl import Workbook

# Read HTML file
with open('amazon_watches.html', 'r') as file:
    html_content = file.read()

# Parse HTML with BeautifulSoup
soup = BeautifulSoup(html_content, 'html.parser')

# Find all div elements with class="a-section a-spacing-base"
div_elements = soup.find_all('div', class_='a-section a-spacing-base')

# Create Excel workbook and select active sheet
workbook = Workbook()
sheet = workbook.active

# Write headers
sheet['A1'] = 'Names'
sheet['B1'] = 'Prices'
sheet['C1'] = 'Reviews'

# Iterate over div elements
for i, div in enumerate(div_elements, start=2):
    # Try to find span with class="a-size-base-plus a-color-base a-text-normal" for Names
    try:
        names = div.find('span', class_='a-size-base-plus a-color-base a-text-normal').text.strip()
    except AttributeError:
        names = ""

    # Try to find span with class="a-price-whole" for Prices
    try:
        prices = div.find('span', class_='a-price-whole').text.strip()
    except AttributeError:
        prices = ""

    # Try to find span with class="a-icon-alt" for Reviews
    try:
        reviews = div.find('span', class_='a-icon-alt').text.strip()
    except AttributeError:
        reviews = ""

    # Write data to Excel
    sheet.cell(row=i, column=1, value=names)
    sheet.cell(row=i, column=2, value=prices)
    sheet.cell(row=i, column=3, value=reviews)

# Save Excel file
workbook.save('amazon_data.xlsx')
