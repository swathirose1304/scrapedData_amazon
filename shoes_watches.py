from bs4 import BeautifulSoup
from openpyxl import Workbook

def extract_data(html_content):
    # Parse HTML with BeautifulSoup
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find all div elements
    div_elements = soup.find_all('div', class_='a-section a-spacing-base')

    # Initialize data list
    data = []

    # Iterate over div elements
    for div in div_elements:
        names = div.find('span', class_='a-size-base-plus a-color-base a-text-normal')
        names = names.text.strip() if names else ""

        prices = div.find('span', class_='a-price-whole')
        prices = prices.text.strip() if prices else ""

        reviews = div.find('span', class_='a-icon-alt')
        reviews = reviews.text.strip() if reviews else ""

        # Append data to the list
        data.append([names, prices, reviews])

    # print(data)
    return  data

# Read HTML files and extract data
shoes_file = 'amazon_shoes.html'
watches_file = 'amazon_watches.html'

with open(shoes_file, 'r') as file:
    shoes_html_content = file.read()

with open(watches_file, 'r') as file:
    watches_html_content = file.read()

shoes_data = extract_data(shoes_html_content)
watches_data = extract_data(watches_html_content)



# Create Excel workbook
workbook = Workbook()

# Create shoes sheet
shoes_sheet = workbook.create_sheet('Shoes', 0)
shoes_sheet.append(['Name', 'Price', 'Reviews'])
for row in shoes_data:
    shoes_sheet.append(row)

# Create watches sheet
watches_sheet = workbook.create_sheet('Watches', 1)
watches_sheet.append(['Name', 'Price', 'Reviews'])
for row in watches_data:
    watches_sheet.append(row)

# Save Excel file
workbook.save('amazon_data.xlsx')
